import subprocess
import gspread
import pandas as pd
from oauth2client.service_account import ServiceAccountCredentials
from flask import Flask, render_template, send_file, request
from datetime import datetime, timedelta
from io import BytesIO
from fpdf import FPDF
from openpyxl import Workbook
from flask import Response

app = Flask(__name__)

# Google Sheets setup
SCOPE = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
CREDENTIALS_FILE = "credentials.json"
SHEET_NAME = "Attendance"

# Authorize Google Sheets access
credentials = ServiceAccountCredentials.from_json_keyfile_name(CREDENTIALS_FILE, SCOPE)
gc = gspread.authorize(credentials)
sheet = gc.open(SHEET_NAME).sheet1

@app.route('/')
def home():
    records = sheet.get_all_records()
    unique_records = {}
    for record in records:
        if record['Roll No'] not in unique_records:
            unique_records[record['Roll No']] = {
                'Roll No': record['Roll No'],
                'Name': record['Name'],
                'Total Attendance': sum(1 for r in records if r['Roll No'] == record['Roll No'])
            }
    return render_template('attendance.html', records=unique_records.values())

@app.route('/start_attendance', methods=['GET'])
def start_attendance():
    try:
        subprocess.Popen(['python', 'main.py'])
        return '', 200
    except Exception as e:
        return str(e), 500

@app.route('/attendance_register')
def attendance_register():
    records = sheet.get_all_records()

    # Extract unique dates from records
    dates = sorted({datetime.strptime(record['Date'], "%Y-%m-%d").date() for record in records if record['Date']})
    today = datetime.now().date()

    # Fill in missing dates
    while dates[-1] < today:
        dates.append(dates[-1] + timedelta(days=1))

    # Initialize attendance data
    attendance_data = {}
    for record in records:
        roll_no = record['Roll No']
        if roll_no not in attendance_data:
            attendance_data[roll_no] = {
                "Name": record['Name'],
                "Attendance": {str(date): "Absent" for date in dates}
            }
        # Mark as present for the specific date
        record_date = datetime.strptime(record['Date'], "%Y-%m-%d").date()
        attendance_data[roll_no]["Attendance"][str(record_date)] = "Present"

    return render_template('attendance_register.html', dates=[str(date) for date in dates], attendance_data=attendance_data)

@app.route('/download_excel')
def download_excel():
    # Fetch all records from the sheet
    records = sheet.get_all_records()

    # Extract all dates and sort them
    from datetime import datetime
    all_dates = sorted(
        {datetime.strptime(record['Date'], "%Y-%m-%d").date() for record in records}
    )

    # Process the attendance data
    attendance_data = {}
    for record in records:
        roll_no = record['Roll No']
        if roll_no not in attendance_data:
            attendance_data[roll_no] = {
                "Name": record['Name'],
                "Attendance": {str(date): "Absent" for date in all_dates}
            }
        attendance_data[roll_no]["Attendance"][record['Date']] = "Present"

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance Register"

    # Add headers
    headers = ["Roll No", "Name"] + [str(date) for date in all_dates]
    ws.append(headers)

    # Add attendance data
    for roll_no, data in attendance_data.items():
        row = [roll_no, data["Name"]] + [
            "P" if data["Attendance"][str(date)] == "Present" else "A" for date in all_dates
        ]
        ws.append(row)

    # Save Excel file to a BytesIO object
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment;filename=attendance-register.xlsx"},
    )

@app.route('/download_pdf')
def download_pdf():
    # Fetch all records from the sheet
    records = sheet.get_all_records()

    # Extract all dates and sort them
    from datetime import datetime
    all_dates = sorted(
        {datetime.strptime(record['Date'], "%Y-%m-%d").date() for record in records}
    )

    # Process the attendance data
    attendance_data = {}
    for record in records:
        roll_no = record['Roll No']
        if roll_no not in attendance_data:
            attendance_data[roll_no] = {
                "Name": record['Name'],
                "Attendance": {str(date): "Absent" for date in all_dates}
            }
        attendance_data[roll_no]["Attendance"][record['Date']] = "Present"

    # Initialize PDF
    pdf = FPDF()
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    # Add title
    pdf.set_font("Arial", size=14, style="B")
    pdf.cell(200, 10, txt="Attendance Register", ln=True, align="C")
    pdf.ln(10)

    # Add table headers
    pdf.set_font("Arial", size=12, style="B")
    pdf.cell(30, 10, "Roll No", border=1)
    pdf.cell(50, 10, "Name", border=1)
    for date in all_dates:
        pdf.cell(20, 10, str(date), border=1)
    pdf.ln()

    # Add table rows
    pdf.set_font("Arial", size=10)
    for roll_no, data in attendance_data.items():
        pdf.cell(30, 10, str(roll_no), border=1)
        pdf.cell(50, 10, data["Name"], border=1)
        for date in all_dates:
            status = data["Attendance"][str(date)]
            pdf.cell(20, 10, "P" if status == "Present" else "A", border=1)
        pdf.ln()

    # Save PDF to a BytesIO object
    output = BytesIO()
    pdf_string = pdf.output(dest='S').encode('latin1')  # Generate PDF as a string
    output.write(pdf_string)
    output.seek(0)

    return send_file(output, as_attachment=True, download_name="attendance-register.pdf", mimetype="application/pdf")



if __name__ == "__main__":
    app.run(debug=True)
